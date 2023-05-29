# ---------------------------------------------------------------------------------------------------------
# foodpicker.py
#   This is the main code of the mealmaster project. Functionalities such as picking a meal plan, altering -
#   it etc. are implemented here. 
# 
# ---------------------------------------------------------------------------------------------------------

import pandas as pd
import random
from flask import Flask, request, session
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
import json
from pymongo import MongoClient
from datetime import datetime
from flask_session import Session
from redis import Redis
import uuid
from datetime import timedelta

app = Flask(__name__)
cors = CORS(app, resources={r"/api/*": {"origins": "*"}})

feedback_file = 'feedback.xlsx'

app.secret_key = 'Xc8zdoe1DWNi3pRLK66DQ7Pz6HRL2Xu3lGpTPco1NyY'
app.config['SESSION_TYPE'] = 'redis'
app.config['SESSION_REDIS'] = Redis(host='localhost', port=6379)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)
app.config['SESSION_COOKIE_HTTPONLY'] = False
Session(app)


@app.before_request
def count_unique_visitors():
    session.permanent = True
    session.modified = True
    if 'session_id' in request.cookies:
        session['session_id'] = request.cookies.get('session_id')
    if 'session_id' not in session:
        session_id = str(uuid.uuid4())
        session['session_id'] = session_id

        client = MongoClient('mongodb://localhost:27017')

        try:
            database = client['theenipandaram']
            collection = database['usercount']
            # Get the current date and time
            current_date = datetime.now()

            # Convert date to string using strftime
            date_string = current_date.strftime('%Y-%m-%d')
            count = collection.count_documents({})
            if count == 0:
                document = {
                    'date': date_string,  # Date in string format
                    'count': 1
                }

                result = collection.insert_one(document)
                datemin_string = current_date.strftime('%Y-%m-%d %H:%M')
                document2 = {
                    'date': datemin_string,
                    'count': 1
                    
                }
                result = collection.insert_one(document2)
            else:
                # Update query to increment the count field for a single document
                filter_query = {
                    'date': date_string
                }

                update_query = {
                    '$inc': {
                        'count': 1
                    }
                }

                result = collection.update_one(filter_query, update_query)
                if result.modified_count==0:
                    document = {
                        'date': date_string,  # Date in string format
                        'count': 1
                    }

                    result = collection.insert_one(document)
                    datemin_string = current_date.strftime('%Y-%m-%d %H:%M')
                    document2 = {
                        'date': datemin_string,
                        'count': 1
                        
                    }
                    result = collection.insert_one(document2)
                else:
                    datemin_string = current_date.strftime('%Y-%m-%d %H:%M')
                    filter_query2 = {
                        'date': datemin_string
                    }

                    update_query2 = {
                        '$inc': {
                            'count': 1
                        }
                    }
                    result = collection.update_one(filter_query2, update_query2)
        finally:
            client.close()

# Set the session ID cookie after each request
@app.after_request
def set_session_id_cookie(response):
    if 'session_id' in session and 'session_id' not in request.cookies:
        response.set_cookie('session_id', session['session_id'], max_age=86400, httponly=False)  # Set cookie validity to one day (86400 seconds)
    return response

#     # Retrieve the custom session identifier from the cookie
@app.before_request
def retrieve_session_id():
    if 'session_id' in request.cookies:
        session['session_id'] = request.cookies.get('session_id')

def insert_update_user_count():
    # Replace with your MongoDB connection string
    client = MongoClient('mongodb://localhost:27017')

    try:
        database = client['theenipandaram']
        collection = database['usercount']
        # Get the current date and time
        current_date = datetime.now()

        # Convert date to string using strftime
        date_string = current_date.strftime('%Y-%m-%d')
        count = collection.count_documents({})
        if count == 0:

            document = {
                'date': date_string,  # Date in string format
                'count': 1
            }

            result = collection.insert_one(document)
            print('Document inserted with ID:', result.inserted_id)
        else:
            # Update query to increment the count field for a single document
            filter_query = {
                'date': date_string
            }

            update_query = {
                '$inc': {
                    'count': 1
                }
            }

            result = collection.update_one(filter_query, update_query)

    finally:
        client.close()


def pick_sidedish(maindish, dishesdb):
    maindishdf = dishesdb[(dishesdb["name"] == maindish)] # This should result in a single row dataframe
    if maindishdf.iloc[0]["sub type"] in ["Complete meal", "unspecified"]:
        return(None)

    if maindishdf.iloc[0]["only legitimate side dishes"] != "":
        legit_sidedish_names = maindishdf.iloc[0]["only legitimate side dishes"].split(", ")
        sidedishesdf = dishesdb[dishesdb["name"].isin(legit_sidedish_names)]
        sidedishesdf.reset_index(drop=True, inplace=True)
        rowval = random.randint(0,sidedishesdf.shape[0]-1)
        sidedish_series = sidedishesdf.loc[rowval]
    else:
        if maindishdf.iloc[0]["sub type"] in ["pongal", "dhida-phalar"]:
            sidedishesdf = dishesdb.query('`sub type` == "chutney"')
        
        elif maindishdf.iloc[0]["sub type"] == "chapathi":
            sidedishesdf = dishesdb.query('`sub type` == "subzi"')
        
        elif maindishdf.iloc[0]["sub type"] == "sambar":
            sidedishesdf = dishesdb.query('`sub type` == ["ambad", "roast", "Vadai"]')
        
        elif maindishdf.iloc[0]["sub type"] == "pilchar":
            sidedishesdf = dishesdb.query('`sub type` == ["poriyal", "kutkiri", "roast", "Vadai"]')
        
        elif maindishdf.iloc[0]["sub type"] == "omty":
            sidedishesdf = dishesdb.query('`sub type` == ["poriyal", "kutkiri", "roast", "Vadai"]')
        
        else:
            sidedishesdf = dishesdb.query('type == "side dish"') 
 
        sidedishesdf.reset_index(drop=True, inplace=True) 
        
        plan_done = False
        rowval = random.randint(0,sidedishesdf.shape[0]-1)
        tries = 0
        bad_sidedishes = maindishdf.iloc[0]["illegitimate side dishes"].split(", ")
        while plan_done == False:
            # Make sure we haven't selected an illegitimate side dish
            if sidedishesdf.loc[rowval, "name"] in bad_sidedishes:
                tries = tries + 1
                if tries <= 3: # Arbitrary choice of max tries. This is to avoid infinite loops
                    rowval = random.randint(0,sidedishesdf.shape[0]-1)
                    continue

            plan_done = True

        sidedish_series = sidedishesdf.loc[rowval]

    sidedish = (sidedish_series.drop(labels=["diet", "time", "only legitimate side dishes", "illegitimate side dishes"])).to_dict()
    return(sidedish) 

def plansinglemeal(dishesdb, mealtime = "lunch"):
    if mealtime == "breakfast":
        subdf = dishesdb.query('time == ["breakfast", "breakfast or dinner"] and type == "main dish"')
    elif mealtime == "dinner":
        subdf = dishesdb.query('time == ["dinner", "breakfast or dinner", "lunch or dinner"] and type == "main dish"')
    else: # Assume mealtime is "lunch"
        subdf = dishesdb.query('time == ["lunch", "lunch or dinner"] and type == "main dish"')

    subdf.reset_index(drop=True, inplace=True)
    items = []
    rowval = random.randint(0,subdf.shape[0]-1)
    maindish_name = subdf.loc[rowval, "name"]
    sidedish = pick_sidedish(maindish = maindish_name, dishesdb = dishesdb)
    if sidedish == None:
        items.append(
                    {
                        "name": maindish_name,
                        "type": "main dish",
                        "sub type" : subdf.loc[rowval, "sub type"],
                        "dosha" : subdf.loc[rowval, "dosha"],
                        "anti dosha": subdf.loc[rowval, "anti dosha"]
                    }
                )
    else:
        items.append({
                        "name": maindish_name,
                        "type": "main dish",
                        "sub type" : subdf.loc[rowval, "sub type"],
                        "dosha" : subdf.loc[rowval, "dosha"],
                        "anti dosha": subdf.loc[rowval, "anti dosha"]
                    })
        items.append(sidedish)

    if mealtime == "lunch":
        if items[0]["sub type"] != "Complete meal": # If we have chosen a complete meal already, then we are done
            if items[0]["sub type"] != "pilchar":
                subsubdf = subdf[(subdf["sub type"] == "pilchar") & (subdf["sub type"] != "Complete meal")]
            else:
                subsubdf = subdf[(subdf["sub type"] != "pilchar") & (subdf["sub type"] != "Complete meal")]

            subsubdf.reset_index(drop=True, inplace=True)
            rowval = random.randint(0,subsubdf.shape[0]-1)
            maindish_name = subsubdf.loc[rowval]["name"]

            plan_done = False
            tries = 0
            sidedish = pick_sidedish(maindish = maindish_name, dishesdb = dishesdb)
            while plan_done == False:
                if sidedish["sub type"] == items[1]["sub type"]:
                    tries = tries + 1
                    if tries <= 3:
                        sidedish = pick_sidedish(maindish = maindish_name, dishesdb = dishesdb)
                        continue

                plan_done = True

            items.append({
                            "name": maindish_name,
                            "type": "main dish",
                            "sub type" : subsubdf.loc[rowval, "sub type"],
                            "dosha" : subsubdf.loc[rowval, "dosha"],
                            "anti dosha": subsubdf.loc[rowval, "anti dosha"]
                        })
            items.append(sidedish)



    meal = {
            "time":mealtime,
            "items":items
           }

    return(meal)



# planmeals-------------------------------------------------------------------------------------------------------  
# Creates a meal plan based on dietary preferences of the user and how many meals he wishes the software to plan
# for.
#
# Arguments:
#   1. diet*: "vegetarian" or "nonvegetarian" 
#   2. planfor: "meal" or "day" or "week"
#   3. mealtime: "breakfast" or "lunch" or "dinner" or None. This cannot be set to None if "planfor" variable is
#                "meal"
#   4. request_type*: "new" or "change"
#   5. change_item: A string mentioning which item needs to be changed. Ex: "day1, meal2, Yellow poosani sambar"
#                   This argument is mandatory only if request_type is set to "change"
#   6. mealplan: The original mealplan json. This argument is mandatory only if request_type is set to "change"  
#
# Return value:
#
#   mealplan: A json variable as described below (with an example).
#   {
#       "day1": {
#           "meal1": {
#               "time": "breakfast",
#               "items": [
#                   {
#                       "name": "Mudha pongal",
#                       "type": "main dish", 
#                       "sub type" : "pongal",
#                       "dosha" : "",
#                       "anti dosha": ""
#                   },
#                   {
#                       "name": "Vangi chutney",
#                       "type": "side dish",
#                       "sub type" : "chutney",
#                       "dosha" : "",
#                       "anti dosha": ""
#                   }
#               ]
#           }
#           "meal2": {
#               "time": "lunch", 
#               "items": [
#                   {
#                       "name": "Yellow poosani sambar",
#                       "type": "main dish",
#                       "sub type" : "sambar",
#                       "dosha" : "vata",
#                       "anti dosha": ""
#                   },
#                   {
#                       "name": "Butter beans ambad",
#                       "type": "side dish",
#                       "sub type" : "ambad",
#                       "dosha" : "vata",
#                       "anti dosha": "unknown"
#                   },
#                   {
#                       "name": "Thikka pilchar",
#                       "type": "main dish",
#                       "sub type" : "pilchar",
#                       "dosha" : "",
#                       "anti dosha": ""
#                   },
#                   {
#                       "name": "Carrot muttakos poriyal",
#                       "type": "side dish",
#                       "sub type" : "poriyal",
#                       "dosha" : "",
#                       "anti dosha": ""
#                   },
#               ]
#           }
#           "meal3": {...},
#           "meal4": {...}
#       },
#   day2: {...},
#   day3: {...}
#   }
#
# Implementation methodology:
#   Since food we eat are associated with human preferences, which to some level is not based on logic, no 
# conventional computer program can produce a meal plan that is flawless from a human's perspective. For ex. We
# might like eating pongal and some variety rices with sambar and chutney, but some others with raita. There are 
# many such unwritten rules. If we try to implement them all in a conventional computer program, the code will 
# become intractable. Instead, we will try to implement some high level rules and then let the users tell us -
# either explicitly by clicking a button in the app that says, "I like this plan" or implicitly when they use the
# retry icon to ask for an alternative suggestion - what combinations of main dishes and side dishes work, or at 
# a higher level, what combinations of breakfast, lunch, dinner work. We will collect the user data and use AI to
# make our suggestions better. For now, this function implements only non AI part - i.e., the part that uses some
# high level food-picking rules. Furthermore, it will only consider the rules for a Tamil audience (in the interest
# of taking it quickly to the market).
#
#   We will use a database called "Dishes_Database" which contains some 200 food items along with their 
# attributes such as whether it is a breakfast/lunch/dinner item, its type (pongal chutney, complete meal etc.), 
# and more.   
#   
#   If the user asks us to give him/her a suggestion for a single meal, we will simply pick randomly an item
# suitable for the time of the day and ensure that the main dish - side dish combinations are legitimate based on 
# some global rules (don't combine chapathi with chutneys, omty with ambads etc.) and also information about
# legitimate combinations explicitly mentioned in the database. 
#
#   If the user asks us to suggests meals for the whole day, we will use the same logic as in the paragrah above,
# but also ensure that we don't accidentally pick the same meal for, say, breakfast and dinner.
#
#   If we are tasked with picking meals for an entire week (5 days), we will use the logic we used in the above two
# paragraphs but additionally, we will use the following:
#       1. We assume that idlis and dosas are quintessial to a Tamil diet. So we will ensure that there is at least
#       one idli and two dosas among a week's breakfasts and dinners taken together. We can do this by simply choosing
#       three slots randomly in breakfasts,dinners and reserving them for one idli means and two dosa meals.
#       2. We have to ensure that idli always comes before dosas. There are reasons why this isn't necessarily the 
#       right choice: People may start with the previous week's batter and make dosas and make a new batter for idli
#       later on. However, we will assume that this is an exception and ignore it. It makes the code simpler. 
#       Furthermore, the user can always use the retry symbol to change the dishes if he/she wishes. So it is not a
#       big deal. 
#        
# ---------------------------------------------------------------------------------------------------------------
    
def changemeal(diet, planfor, day, mealtime, request_type, change_item, meal_plan):
    dishesdb = pd.read_excel("Dishes_Database.xlsx", dtype=str).fillna("")

    if diet == "vegetarian":
        dishesdb.drop(dishesdb[dishesdb['diet'] == "nonvegetarian"].index, inplace = True)

    time_to_meal = {"breakfast":"meal1", "lunch":"meal2", "dinner":"meal3"} # Unnecessary step if instead of mealtime, the meal number (meal1, meal2 etc.) is sent
    for item_idx, item in enumerate(meal_plan[day][time_to_meal[mealtime]]["items"]): # Find the item that needs to be changed
        if item["name"] == change_item:
            break

    if item["type"] == "main dish":
        if mealtime == "breakfast":
            legit_mealtimes = ["breakfast", "breakfast or dinner"]
            if item["sub type"] in ["pongal", "dhida-phalar", "chapathi"]:
                legit_subtypes =  ["pongal", "dhida-phalar", "chapathi"]
            else:
                legit_subtypes =  ["pongal", "dhida-phalar", "chapathi", "Complete meal"]
        elif mealtime == "dinner":
            legit_mealtimes = ["dinner", "lunch or dinner", "breakfast or dinner"]
            if item["sub type"] in ["pongal", "dhida-phalar", "chapathi"]:
                legit_subtypes = ["pongal", "dhida-phalar", "chapathi"]
            else:
                legit_subtypes = ["pongal", "dhida-phalar", "chapathi", "Complete meal"]
        else: # TODO: We arent considering snacks for now
            legit_mealtimes = ["lunch", "lunch or dinner"]
            if item["sub type"] in ["sambar", "omty"]:
                legit_subtypes =  ["sambar", "omty"]
            elif item["sub type"] == "pilchar":
                legit_subtypes =  ["pilchar"]
            else:
                legit_subtypes = ["sambar", "omty", "Complete meal"]

    else:
        if mealtime == "breakfast":
            legit_mealtimes = ["breakfast", "breakfast or dinner"]
            if item["sub type"] == "chutney" or "unspecified":
                legit_subtypes = ["chutney"]
            elif item["sub type"] == "subzi":
                legit_subtypes = ["subzi"]
        elif mealtime == "dinner":
            legit_mealtimes = ["dinner", "lunch or dinner", "breakfast or dinner"]
            if item["sub type"] == "chutney" or "unspecified":
                legit_subtypes = ["chutney"]
            elif item["sub type"] == "subzi":
                legit_subtypes = ["subzi"]
        else:
            legit_mealtimes = ["lunch", "lunch or dinner"]
            legit_subtypes = ["ambad", "roast", "Vadai", "poriyal", "kutkiri", "roast", "Vadai"]
    
    subdf = dishesdb[(dishesdb["time"].isin(legit_mealtimes)) & (dishesdb["sub type"].isin(legit_subtypes)) &(dishesdb["name"] != item["name"])]
    try:
        new_dish_series = subdf.iloc[random.randint(0, subdf.shape[0]-1)]
    except Exception as err:
        print("Some exception ", err)
    new_dish = (new_dish_series.drop(labels=["diet", "time", "only legitimate side dishes", "illegitimate side dishes"])).to_dict()
    meal_plan[day][time_to_meal[mealtime]]["items"][item_idx] = new_dish
    

    # If original item was a Complete meal and we selected a omty/sambar/pilchar in its place, then we need to choose a sidedish
    if item["sub type"] == "Complete meal":
        if new_dish["sub type"] in ["omty", "sambar", "pilchar", "pongal", "dhida-phalar", "chapathi"]:
            sidedish = pick_sidedish(maindish=new_dish["name"], dishesdb=dishesdb)
            meal_plan[day][time_to_meal[mealtime]]["items"].append(sidedish)

    return(meal_plan)

@app.route('/api/v1.0/feedback', methods=['POST'])
def feedback():
    name = request.form.get('name')
    whatsapp_number = request.form.get('whatsappNumber')
    email = request.form.get('email')
    feedback = request.form.get('feedback')

    # Load existing workbook
    try:
        workbook = load_workbook(feedback_file)
        sheet = workbook.active
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'WhatsApp Number', 'Email', 'Feedback'])

    # Append form data to the workbook
    sheet.append([name, whatsapp_number, email, feedback])
    workbook.save(feedback_file)
    return 'Form data submitted successfully!'

@app.route('/api/v1.0')
# @cross_origin()
def planmeals_api():
    args = request.args
    return planmeals(args.get("diet"), args.get("planfor"), args.get("day"), args.get("mealtime"), args.get("request_type"), args.get("change_item"), args.get("meal_plan"))

@app.route('/api/v1.0', methods=['POST'])
def planmeals_api_json():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        data = json.loads(request.data)
        return changemeal(diet=data["diet"], planfor=data["planfor"], day=data["day"],mealtime=data["mealtime"], request_type="change", change_item=data["change_item"], meal_plan=json.loads(data["meal_plan"]))
    else:
        return 'Content-Type not supported!'


def planmeals(diet="vegetarian", planfor="day", day="day", mealtime="lunch", request_type="new", change_item=None, meal_plan=None):
    dishesdb = pd.read_excel("Dishes_Database.xlsx", dtype=str).fillna("")

    if diet == "vegetarian":
        dishesdb.drop(dishesdb[dishesdb['diet'] == "nonvegetarian"].index, inplace = True)

    if planfor == "meal":
        meal = plansinglemeal(mealtime = mealtime, dishesdb = dishesdb)
        mealplan = {"day":{"meal":meal}}

    elif planfor == "day":
        meal1 = plansinglemeal(mealtime = "breakfast", dishesdb = dishesdb)
        meal2 = plansinglemeal(mealtime = "lunch", dishesdb = dishesdb)
        
        plan_done = False   # This variable is used to tell us when to stop iterating on 
        meal3 = plansinglemeal(mealtime="dinner", dishesdb=dishesdb)
        tries = 0
        while plan_done == False:
            if meal1["items"][0]["sub type"] == meal3["items"][0]["sub type"]:  # Assuming items[0] will always be the main dish
                                                                                # We don't want two dishes of the same sub type,
                                                                                # ex. pongal, to be selected for both breakfast 
                                                                                # and dinner
                tries = tries+1
                if tries <= 3:
                    meal3 = plansinglemeal(mealtime="dinner", dishesdb=dishesdb)
                    continue

            if len(meal1["items"]) == 2 and len(meal3["items"]) == 2 :
                if meal1["items"][1]["name"] == meal3["items"][1]["name"]:  # Assuming items[1] will always be the side dish
                                                                            # We don't want the same side dish for morning and night
                   tries = tries + 1
                   if tries <= 3:
                       meal3 = plansinglemeal(mealtime="dinner", dishesdb=dishesdb)
                       continue

            plan_done = True
        
        mealplan = {"day":{"meal1":meal1, "meal2":meal2, "meal3":meal3}}

    else:   # Assume the plan request is for a week
        # Let us first select breakfast and dinner
        bf_subtypes = ["pongal", "pongal", "pongal", "dhida-phalar", "dhida-phalar", "dhida-phalar", "chapathi",
                       "chapathi", "Complete meal"]
        dinner_subtypes = ["pongal", "pongal", "dhida-phalar", "dhida-phalar", "dhida-phalar", "chapathi", "chapathi",
                           "Complete meal"]
       
        # Ensure breakfast items have enough diversity
        step_done = False
        bf_choices = random.choices(bf_subtypes, k=5)
        tries = 0
        while step_done == False:
            if bf_choices.count("pongal") > 3 or\
                bf_choices.count("dhida-phalar") > 3 or\
                bf_choices.count("chapati") > 2:

                tries = tries+1
                if tries <= 3:
                    bf_choices = random.choices(bf_subtypes, k=5)
                    continue

            step_done = True

        plan_done = False
        dinner_choices = random.choices(dinner_subtypes, k=5)
        tries = 0
        while plan_done == False:
            # Ensure dinner items have enough diversity
            step_done = False
            while step_done == False:
                if dinner_choices.count("pongal") > 3 or\
                    dinner_choices.count("dhida-phalar") > 3 or\
                    dinner_choices.count("chapati") > 2:

                    tries = tries + 1
                    if tries <= 6:
                        dinner_choices = random.choices(dinner_subtypes, k=5)
                        continue

                step_done = True
           
            # Check if there is enough diversity among breakfast and dinner items combined 
            temp_list = []
            temp_list.extend(bf_choices)
            temp_list.extend(dinner_choices)
            if temp_list.count("pongal") > 5 or\
                temp_list.count("dhida-phalar") > 5 or\
                temp_list.count("chapati") > 3:

                tries = tries + 1
                if tries <= 6:
                    dinner_choices = random.choices(dinner_subtypes, k=5)
                continue

            plan_done = True
            
        # Rearrange the dinner dishes so that we don't get two chapati meals on the same day, two pongals on the same day and so on
        no_of_tries = 5
        step_done = False
        tries = 0
        while step_done == False:
            matches = 0
            for i in range(len(bf_choices)):
                if bf_choices[i] == dinner_choices[i]:
                    matches = matches + 1
                
            if (matches > 2):
                tries = tries+1
                if tries <= 3:
                    random.shuffle(dinner_choices)
                    continue

            step_done = True
                
        # We are done with choosing breakfast and dinner. Lets choose lunch meals now.
        lunch_subtypes = ["omty", "omty", "omty", "sambar", "sambar", "sambar", "Complete meal"]    # omty and sambar will always be accompanied by 
                                                                                                    # a pilchar. 

        step_done = False
        lunch_choices = random.choices(lunch_subtypes, k=5)
        tries = 0
        while step_done == False:
            # Ensure diversity of lunch items
            if lunch_choices.count("omty") > 3 or\
                lunch_choices.count("sambar") > 3 or\
                lunch_choices.count("Complete meal") > 2:

                tries = tries+1
                if tries <= 3:
                    lunch_choices = random.choices(lunch_subtypes, k=5)
                    continue

            step_done = True
    
        
        # Select main dishes for all meals
        bf_df = dishesdb.query('time == ["breakfast", "breakfast or dinner"] and type == "main dish"')
        dinner_df = dishesdb.query('time == ["dinner", "breakfast or dinner", "lunch or dinner"] and type == "main dish"')
        lunch_df = dishesdb.query('time == ["lunch", "lunch or dinner"] and type == "main dish"')
        
        # Choose actual breakfast dishes
        bf_dishes = []
        prev_bf_maindishes = []
        prev_bf_sidedishes = []
        for dishsubtype in bf_choices:
            subdf = bf_df[(bf_df["sub type"] == dishsubtype)]
            subdf.reset_index(drop=True, inplace=True) 
            
            maindish_chosen = False
            try:
                rowval = random.randint(0, subdf.shape[0]-1)
            except:
                print("subdf was probably empty because bf_choices is empty")
            tries = 0
            while maindish_chosen == False:
                maindish = {
                            "name": subdf.loc[rowval, "name"],
                            "type": subdf.loc[rowval, "type"],
                            "sub type" : subdf.loc[rowval, "sub type"],
                            "dosha" : subdf.loc[rowval, "dosha"],
                            "anti dosha": subdf.loc[rowval, "anti dosha"]
                            }
                if maindish["name"] in prev_bf_maindishes:
                    tries = tries+1
                    if tries <= 3:
                        rowval = random.randint(0, subdf.shape[0]-1)
                        continue 
                
                maindish_chosen = True

            prev_bf_maindishes.append(maindish["name"])
            
            sidedish_chosen = False
            sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
            tries = 0
            while sidedish_chosen == False:
                if sidedish == None:
                    sidedish_chosen = True
                    break
                if sidedish["name"] in prev_bf_sidedishes:
                    tries = tries+1
                    if tries <= 3:
                        sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                        continue

                sidedish_chosen = True
            
            if sidedish != None:
                prev_bf_sidedishes.append(sidedish["name"])
                bf_dishes.append([maindish, sidedish])
            else:
                bf_dishes.append([maindish,{}])

        # Choose actual dinner dishes
        dinner_dishes = []
        prev_dinner_maindishes = []
        prev_dinner_sidedishes = []
        for i, dishsubtype in enumerate(dinner_choices):
            subdf = dinner_df[(dinner_df["sub type"] == dishsubtype)]
            subdf.reset_index(drop=True, inplace=True) 
            
            maindish_chosen = False
            try:
                rowval = random.randint(0, subdf.shape[0]-1)
            except:
                print("subdf was probably empty because dinner_choices is empty")
            tries = 0
            while maindish_chosen == False:
                maindish = {
                            "name": subdf.loc[rowval, "name"],
                            "type": subdf.loc[rowval, "type"],
                            "sub type" : subdf.loc[rowval, "sub type"],
                            "dosha" : subdf.loc[rowval, "dosha"],
                            "anti dosha": subdf.loc[rowval, "anti dosha"]
                            }
                if maindish["name"] in prev_dinner_maindishes:
                    tries = tries+1
                    if tries <= 3:
                        rowval = random.randint(0, subdf.shape[0]-1)
                        continue 
                
                maindish_chosen = True

            prev_dinner_maindishes.append(maindish["name"])
            
            sidedish_chosen = False
            sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
            tries = 0
            while sidedish_chosen == False:
                if sidedish == None:
                    sidedish_chosen = True
                    break
                if sidedish["name"] in prev_dinner_sidedishes:
                    tries = tries+1
                    if tries <= 3:
                        sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                        continue
                if bf_dishes[i][1] != {}:
                    if (sidedish["name"] == bf_dishes[i][1]["name"]):
                        tries = tries+1
                        if tries <= 3:
                            sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                            continue

                sidedish_chosen = True
            
            if sidedish != None:
                prev_dinner_sidedishes.append(sidedish["name"])
                dinner_dishes.append([maindish, sidedish])
            else:
                dinner_dishes.append([maindish,{}])
                



        # Choose actual lunch dishes
        lunch_dishes = []
        prev_lunch_maindishes = []
        prev_lunch_sidedishes = []
        for i, dishsubtype in enumerate(lunch_choices):
            subdf = lunch_df[(lunch_df["sub type"] == dishsubtype)]
            subdf.reset_index(drop=True, inplace=True) 
            
            maindish_chosen = False
            try:
                rowval = random.randint(0, subdf.shape[0]-1)
            except:
                print("subdf was probably empty because lunch_choices is empty")
            tries = 0
            while maindish_chosen == False:
                maindish = {
                            "name": subdf.loc[rowval, "name"],
                            "type": subdf.loc[rowval, "type"],
                            "sub type" : subdf.loc[rowval, "sub type"],
                            "dosha" : subdf.loc[rowval, "dosha"],
                            "anti dosha": subdf.loc[rowval, "anti dosha"]
                            }
                if maindish["name"] in prev_lunch_maindishes:
                    tries = tries+1
                    if tries <= 3:
                        rowval = random.randint(0, subdf.shape[0]-1)
                        continue 
                
                maindish_chosen = True

            prev_lunch_maindishes.append(maindish["name"])
            
            sidedish_chosen = False
            sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
            tries = 0
            while sidedish_chosen == False:
                if sidedish == None:
                    sidedish_chosen = True
                    break
                if sidedish["name"] in prev_lunch_sidedishes:
                    tries = tries+1
                    if tries <= 3:
                        sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                        continue

                sidedish_chosen = True
            
            if sidedish != None:
                prev_lunch_sidedishes.append(sidedish["name"])
                lunch_dishes.append([maindish, sidedish])
            else:
                lunch_dishes.append([maindish,{}])
            
            if maindish["sub type"] in ["omty", "sambar"]:  # If the dish we just selected is an omty or a sambar, then we need
                                                            # a pilchar to accompany it
                subdf = lunch_df[(lunch_df["sub type"] == "pilchar")]
                subdf.reset_index(drop=True, inplace=True) 
                
                rowval = random.randint(0, subdf.shape[0]-1)
                maindish = {
                            "name": subdf.loc[rowval, "name"],
                            "type": subdf.loc[rowval, "type"],
                            "sub type" : subdf.loc[rowval, "sub type"],
                            "dosha" : subdf.loc[rowval, "dosha"],
                            "anti dosha": subdf.loc[rowval, "anti dosha"]
                            }
            
                sidedish_chosen = False
                sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                tries = 0
                while sidedish_chosen == False:
                    if sidedish == None:
                        sidedish_chosen = True
                        break
                    if sidedish["name"] in prev_lunch_sidedishes:
                        tries = tries+1
                        if tries <= 3:
                            sidedish = pick_sidedish(maindish = maindish["name"], dishesdb = dishesdb)
                            continue

                    sidedish_chosen = True
                
                if sidedish != None:
                    prev_lunch_sidedishes.append(sidedish["name"])
                    lunch_dishes.append([maindish, sidedish])
                else:
                    lunch_dishes.append([maindish,{}])
                
            else:
                lunch_dishes.append([{}, {}])

        # Build the mealplan json from all the dishes selected in the above steps.
        days = []
        for day in range(5):
            meal1_items = bf_dishes[day]
            meal2_items = lunch_dishes[day*2]
            meal2_items.extend(lunch_dishes[day*2 + 1])
            meal3_items = dinner_dishes[day]

            days.append({
                            "meal1": {"time":"breakfast", "items": meal1_items},
                            "meal2": {"time":"lunch", "items": meal2_items},
                            "meal3": {"time":"dinner", "items": meal3_items}
                        })
        
        mealplan = {
                    "day1": days[0],
                    "day2": days[1],
                    "day3": days[2],
                    "day4": days[3],
                    "day5": days[4],
                    }

    return(mealplan)


# mealplan = planmeals(planfor="day")
# print(json.dumps(mealplan, indent=4))
if __name__ == '__main__':
    app.run('0.0.0.0', port=5000)
